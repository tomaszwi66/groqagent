# ─────────────────────────────────────────
# ENCODING - MUST BE FIRST
# ─────────────────────────────────────────
import sys
import os
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
os.environ["PYTHONIOENCODING"] = "utf-8"

# ─────────────────────────────────────────
# STANDARD IMPORTS
# ─────────────────────────────────────────
import json
import time
import urllib.request
from html.parser import HTMLParser
from datetime import datetime, date

# ─────────────────────────────────────────
# GROQ - AI CLIENT
# ─────────────────────────────────────────
from groq import Groq

API_KEY = ""  # ← PASTE YOUR API KEY HERE
client = Groq(api_key=API_KEY)

# ─────────────────────────────────────────
# MODELS - DUAL STRATEGY
# ─────────────────────────────────────────
MODEL_FAST  = "llama-3.1-8b-instant"       # 14,400 req/day - simple tasks
MODEL_SMART = "llama-3.3-70b-versatile"    # 1,000 req/day  - complex tasks

_smart_calls_today = 0
_smart_calls_date  = date.today()
_MAX_SMART_CALLS   = 800  # safety buffer


def choose_model(user_message: str) -> str:
    """Automatically selects a model based on task complexity."""
    global _smart_calls_today, _smart_calls_date

    # Reset counter at midnight
    if date.today() != _smart_calls_date:
        _smart_calls_today = 0
        _smart_calls_date  = date.today()

    msg_lower = user_message.lower()

    simple_keywords = [
        "open", "click", "type", "screenshot", "save", "read",
        "scroll", "wait", "close", "show", "list", "go to", "navigate",
    ]
    complex_keywords = [
        "analyz", "compare", "strategy", "plan", "summarize", "summary",
        "report", "insight", "optim", "explain", "why", "create", "design",
        "budget", "calcul", "step by step", "excel", "chart", "html", "table",
    ]

    is_simple  = any(kw in msg_lower for kw in simple_keywords)
    is_complex = any(kw in msg_lower for kw in complex_keywords)

    # Long prompts (>200 chars) are always treated as complex
    if len(user_message) > 200:
        is_complex = True

    if is_complex and _smart_calls_today < _MAX_SMART_CALLS:
        _smart_calls_today += 1
        return MODEL_SMART
    elif is_simple and not is_complex:
        return MODEL_FAST
    elif _smart_calls_today < _MAX_SMART_CALLS:
        _smart_calls_today += 1
        return MODEL_SMART
    else:
        return MODEL_FAST


# ─────────────────────────────────────────
# PLAYWRIGHT
# ─────────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("[⚠️ Playwright not available. Run: pip install playwright && playwright install chromium]")

# ─────────────────────────────────────────
# OPENPYXL (EXCEL)
# ─────────────────────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[⚠️ openpyxl not available. Run: pip install openpyxl]")

# ─────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")

# Playwright - single instance per session
_playwright = None
_browser    = None
_page       = None

# ─────────────────────────────────────────
# HELPER - PATH NORMALIZATION
# ─────────────────────────────────────────
def fix_path(path: str) -> str:
    return (path
            .replace("~/Desktop",  DESKTOP)
            .replace("/desktop",   DESKTOP)
            .replace("~/desktop",  DESKTOP))


# ─────────────────────────────────────────
# BROWSER
# ─────────────────────────────────────────
def get_page():
    global _playwright, _browser, _page
    if _page is None:
        _playwright = sync_playwright().start()
        _browser    = _playwright.chromium.launch(headless=False, slow_mo=100)
        _page       = _browser.new_page()
        _page.set_viewport_size({"width": 1280, "height": 800})
    return _page


def close_browser():
    global _playwright, _browser, _page
    if _browser:
        try: _browser.close()
        except: pass
    if _playwright:
        try: _playwright.stop()
        except: pass
    _page = _browser = _playwright = None


def browser_goto(url: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        if not url.startswith("http"):
            url = "https://" + url
        page = get_page()
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        return f"Opened: {url} | Title: {page.title()}"
    except Exception as e:
        return f"Navigation error: {e}"


def browser_click(selector: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        clicked = False
        for strategy in [
            lambda: page.get_by_text(selector, exact=False).first.click(timeout=3000),
            lambda: page.get_by_role("button", name=selector).first.click(timeout=3000),
            lambda: page.get_by_role("link",   name=selector).first.click(timeout=3000),
            lambda: page.click(selector, timeout=3000),
        ]:
            try:
                strategy()
                clicked = True
                break
            except: pass
        if not clicked:
            return f"Element not found: {selector}"
        page.wait_for_load_state("domcontentloaded")
        return f"Clicked: {selector}"
    except Exception as e:
        return f"Click error: {e}"


def browser_type(selector: str, text: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        filled = False
        for strategy in [
            lambda: page.get_by_placeholder(selector).first.fill(text, timeout=3000),
            lambda: page.get_by_label(selector).first.fill(text, timeout=3000),
            lambda: page.get_by_role("textbox", name=selector).first.fill(text, timeout=3000),
            lambda: page.get_by_role("searchbox").first.fill(text, timeout=3000),
            lambda: page.fill(selector, text, timeout=3000),
        ]:
            try:
                strategy()
                filled = True
                break
            except: pass
        if not filled:
            return f"Input field not found: {selector}"
        return f"Typed '{text}' into: {selector}"
    except Exception as e:
        return f"Type error: {e}"


def browser_get_text() -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        text = page.inner_text("body")
        if len(text) > 6000:
            text = text[:6000] + "\n[... truncated to 6000 chars]"
        return text
    except Exception as e:
        return f"Get text error: {e}"


def browser_screenshot(path: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    path = fix_path(path)
    if not path.endswith(".png"):
        path += ".png"
    try:
        page = get_page()
        dir_name = os.path.dirname(path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        page.screenshot(path=path, full_page=True)
        return f"Screenshot saved: {path}"
    except Exception as e:
        return f"Screenshot error: {e}"


def browser_get_links() -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        links = page.eval_on_selector_all(
            "a[href]",
            "els => els.map(e => ({text: e.innerText.trim(), href: e.href}))"
            ".filter(l => l.text && l.href)"
        )
        result = [f"{l['text'][:60]} -> {l['href']}" for l in links[:40]]
        return "\n".join(result) if result else "No links found."
    except Exception as e:
        return f"Get links error: {e}"


def browser_scroll(direction: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        key_map = {"down": "PageDown", "up": "PageUp", "top": "Home", "bottom": "End"}
        page.keyboard.press(key_map.get(direction, "PageDown"))
        time.sleep(0.3)
        return f"Scrolled: {direction}"
    except Exception as e:
        return f"Scroll error: {e}"


def browser_press_key(key: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        page.keyboard.press(key)
        time.sleep(0.5)
        return f"Pressed: {key}"
    except Exception as e:
        return f"Key error: {e}"


def browser_select_option(selector: str, value: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        try:
            page.select_option(selector, label=value, timeout=5000)
        except:
            page.select_option(selector, value=value, timeout=5000)
        return f"Selected '{value}' in: {selector}"
    except Exception as e:
        return f"Select error: {e}"


def browser_wait(seconds) -> str:
    try:
        sec = min(float(seconds), 30)
        time.sleep(sec)
        return f"Waited {sec}s"
    except Exception as e:
        return f"Wait error: {e}"


def browser_current_url() -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        return f"URL: {page.url} | Title: {page.title()}"
    except Exception as e:
        return f"URL error: {e}"


def browser_go_back() -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        page.go_back(wait_until="domcontentloaded", timeout=10000)
        return f"Went back. URL: {page.url}"
    except Exception as e:
        return f"Go back error: {e}"


def browser_eval_js(script: str) -> str:
    if not PLAYWRIGHT_AVAILABLE:
        return "Playwright not available."
    try:
        page = get_page()
        result = page.evaluate(script)
        return str(result)[:3000] if result else "OK (no result)"
    except Exception as e:
        return f"JS error: {e}"


# ─────────────────────────────────────────
# WEB (without browser)
# ─────────────────────────────────────────
class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.text = []
        self.skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style", "nav", "footer", "head", "noscript"):
            self.skip = True

    def handle_endtag(self, tag):
        if tag in ("script", "style", "nav", "footer", "head", "noscript"):
            self.skip = False

    def handle_data(self, data):
        if not self.skip:
            s = data.strip()
            if s:
                self.text.append(s)


def read_webpage(url: str) -> str:
    try:
        if not url.startswith("http"):
            url = "https://" + url
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        with urllib.request.urlopen(req, timeout=15) as r:
            html = r.read().decode("utf-8", errors="ignore")
        parser = TextExtractor()
        parser.feed(html)
        text = "\n".join(parser.text)
        if len(text) > 8000:
            text = text[:8000] + "\n[... truncated to 8000 chars]"
        return text or "No content."
    except Exception as e:
        return f"Error: {e}"


# ─────────────────────────────────────────
# FILES
# ─────────────────────────────────────────
def read_file(path: str) -> str:
    path = fix_path(path)
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        if len(content) > 10000:
            content = content[:10000] + "\n[... truncated to 10000 chars]"
        return content
    except UnicodeDecodeError:
        try:
            with open(path, "r", encoding="cp1250") as f:
                return f.read()[:10000]
        except Exception as e2:
            return f"Read error (encoding): {e2}"
    except Exception as e:
        return f"Read error: {e}"


def write_file(path: str, content: str) -> str:
    path = fix_path(path)
    try:
        dir_name = os.path.dirname(path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"Saved: {path} ({len(content)} chars)"
    except Exception as e:
        return f"Write error: {e}"


def list_files(directory: str = ".") -> str:
    directory = fix_path(directory)
    try:
        entries = os.listdir(directory)
        result  = []
        for entry in sorted(entries):
            full = os.path.join(directory, entry)
            if os.path.isdir(full):
                result.append(f"📁 {entry}/")
            else:
                size = os.path.getsize(full)
                if   size < 1024:            size_str = f"{size} B"
                elif size < 1024 * 1024:     size_str = f"{size/1024:.1f} KB"
                else:                        size_str = f"{size/1024/1024:.1f} MB"
                result.append(f"📄 {entry} ({size_str})")
        return "\n".join(result) if result else "Directory is empty."
    except Exception as e:
        return f"Error: {e}"


def open_file(path: str) -> str:
    path = fix_path(path)
    try:
        os.startfile(path)
        return f"Opened: {path}"
    except Exception as e:
        return f"Error: {e}"


def delete_file(path: str) -> str:
    path = fix_path(path)
    try:
        if os.path.isfile(path):
            os.remove(path)
            return f"Deleted file: {path}"
        elif os.path.isdir(path):
            import shutil
            shutil.rmtree(path)
            return f"Deleted folder: {path}"
        else:
            return f"Not found: {path}"
    except Exception as e:
        return f"Delete error: {e}"


def copy_file(src: str, dst: str) -> str:
    src, dst = fix_path(src), fix_path(dst)
    try:
        import shutil
        dir_name = os.path.dirname(dst)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        shutil.copy2(src, dst)
        return f"Copied: {src} -> {dst}"
    except Exception as e:
        return f"Copy error: {e}"


def move_file(src: str, dst: str) -> str:
    src, dst = fix_path(src), fix_path(dst)
    try:
        import shutil
        dir_name = os.path.dirname(dst)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        shutil.move(src, dst)
        return f"Moved: {src} -> {dst}"
    except Exception as e:
        return f"Move error: {e}"


def create_directory(path: str) -> str:
    path = fix_path(path)
    try:
        os.makedirs(path, exist_ok=True)
        return f"Created folder: {path}"
    except Exception as e:
        return f"Error: {e}"


# ─────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────
def _thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def create_excel(path: str, sheets_data: list) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    if not path.endswith(".xlsx"):
        path += ".xlsx"
    try:
        wb = Workbook()
        wb.remove(wb.active)
        for sd in sheets_data:
            ws         = wb.create_sheet(title=sd.get("name", "Sheet1"))
            headers    = sd.get("headers", [])
            rows       = sd.get("rows",    [])
            col_widths = sd.get("col_widths", [])

            if headers:
                ws.append(headers)
                for cell in ws[1]:
                    cell.font      = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill      = PatternFill("solid", fgColor="4472C4")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = _thin_border()

            for row in rows:
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.border    = _thin_border()
                    cell.alignment = Alignment(vertical="center")

            if col_widths:
                for i, w in enumerate(col_widths):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
            else:
                for col in ws.columns:
                    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        dir_name = os.path.dirname(path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)
        wb.save(path)
        return f"Excel '{path}' created ({len(sheets_data)} sheets)."
    except Exception as e:
        return f"Excel create error: {e}"


def read_excel(path: str) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb     = load_workbook(path, data_only=True)
        result = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.append(f"=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    result.append("\t".join(str(c) if c is not None else "" for c in row))
                    row_count += 1
                    if row_count > 100:
                        result.append("[... truncated to 100 rows]")
                        break
        return "\n".join(result) or "File is empty."
    except Exception as e:
        return f"Excel read error: {e}"


def edit_excel_cell(path: str, sheet_name: str, cell: str, value) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        try:
            value = float(value)
            if value == int(value):
                value = int(value)
        except (ValueError, TypeError):
            pass
        ws[cell] = value
        wb.save(path)
        return f"Cell {sheet_name}!{cell} = {value}"
    except Exception as e:
        return f"Cell edit error: {e}"


def add_excel_formula(path: str, sheet_name: str, cell: str, formula: str) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws[cell] = formula
        wb.save(path)
        return f"Formula '{formula}' set in {sheet_name}!{cell}"
    except Exception as e:
        return f"Formula error: {e}"


def add_excel_chart(path: str, sheet_name: str, chart_type: str,
                    data_range: str, title: str, position: str) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb  = load_workbook(path)
        ws  = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ref = Reference(ws, range_string=f"{sheet_name}!{data_range}")
        chart_map   = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        chart_class = chart_map.get(chart_type, BarChart)
        chart        = chart_class()
        chart.title  = title
        chart.style  = 10
        chart.width  = 18
        chart.height = 12
        chart.add_data(ref, titles_from_data=True)
        ws.add_chart(chart, position)
        wb.save(path)
        return f"Chart '{chart_type}' '{title}' added at {position}."
    except Exception as e:
        return f"Chart error: {e}"


def add_excel_sheet(path: str, sheet_name: str) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        wb.save(path)
        return f"Sheet '{sheet_name}' added."
    except Exception as e:
        return f"Add sheet error: {e}"


def excel_add_rows(path: str, sheet_name: str, rows: list) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in rows:
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = _thin_border()
        wb.save(path)
        return f"Added {len(rows)} rows to '{sheet_name}'."
    except Exception as e:
        return f"Add rows error: {e}"


def excel_style_range(path: str, sheet_name: str, cell_range: str,
                      bold: bool = False, bg_color: str = None, font_size: int = None) -> str:
    if not EXCEL_AVAILABLE:
        return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in ws[cell_range]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                if bold or font_size:
                    cell.font = Font(bold=bold, size=font_size or cell.font.size)
                if bg_color:
                    cell.fill = PatternFill("solid", fgColor=bg_color)
        wb.save(path)
        return f"Style applied to {cell_range}."
    except Exception as e:
        return f"Style error: {e}"


# ─────────────────────────────────────────
# SYSTEM COMMANDS
# ─────────────────────────────────────────
def run_command(command: str) -> str:
    try:
        import subprocess
        result = subprocess.run(
            command, shell=True, capture_output=True, text=True,
            timeout=30, encoding="utf-8", errors="ignore"
        )
        output = result.stdout + result.stderr
        if len(output) > 5000:
            output = output[:5000] + "\n[... truncated]"
        return output or "Command executed (no output)."
    except subprocess.TimeoutExpired:
        return "Timeout - command took too long."
    except Exception as e:
        return f"Error: {e}"


# ─────────────────────────────────────────
# TOOLS DEFINITION
# ─────────────────────────────────────────
tools = [
    # ── FILES ───────────────────────────────────────────────────────────────
    {"type": "function", "function": {
        "name": "read_file",
        "description": "Read the contents of a text file (txt, py, html, csv, json, etc.).",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Path to the file"}},
            "required": ["path"]}}},

    {"type": "function", "function": {
        "name": "write_file",
        "description": "Write text content to a file. Creates parent directories if needed.",
        "parameters": {"type": "object", "properties": {
            "path":    {"type": "string"},
            "content": {"type": "string"}},
            "required": ["path", "content"]}}},

    {"type": "function", "function": {
        "name": "list_files",
        "description": "List files and folders in a directory with sizes.",
        "parameters": {"type": "object", "properties": {
            "directory": {"type": "string", "description": "Directory path. Defaults to current directory."}},
            "required": []}}},

    {"type": "function", "function": {
        "name": "open_file",
        "description": "Open a file in its default Windows application.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string"}},
            "required": ["path"]}}},

    {"type": "function", "function": {
        "name": "delete_file",
        "description": "Delete a file or folder.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string"}},
            "required": ["path"]}}},

    {"type": "function", "function": {
        "name": "copy_file",
        "description": "Copy a file to a new location.",
        "parameters": {"type": "object", "properties": {
            "src": {"type": "string"},
            "dst": {"type": "string"}},
            "required": ["src", "dst"]}}},

    {"type": "function", "function": {
        "name": "move_file",
        "description": "Move a file to a new location.",
        "parameters": {"type": "object", "properties": {
            "src": {"type": "string"},
            "dst": {"type": "string"}},
            "required": ["src", "dst"]}}},

    {"type": "function", "function": {
        "name": "create_directory",
        "description": "Create a folder (recursive).",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string"}},
            "required": ["path"]}}},

    # ── BROWSER ─────────────────────────────────────────────────────────────
    {"type": "function", "function": {
        "name": "browser_goto",
        "description": "Open a URL in Chromium browser (Playwright).",
        "parameters": {"type": "object", "properties": {
            "url": {"type": "string"}},
            "required": ["url"]}}},

    {"type": "function", "function": {
        "name": "browser_click",
        "description": "Click an element on the page - provide visible text or CSS selector.",
        "parameters": {"type": "object", "properties": {
            "selector": {"type": "string"}},
            "required": ["selector"]}}},

    {"type": "function", "function": {
        "name": "browser_type",
        "description": "Type text into a form field. Selector can be placeholder, label, or CSS.",
        "parameters": {"type": "object", "properties": {
            "selector": {"type": "string"},
            "text":     {"type": "string"}},
            "required": ["selector", "text"]}}},

    {"type": "function", "function": {
        "name": "browser_get_text",
        "description": "Get all visible text from the current page (max 6000 chars).",
        "parameters": {"type": "object", "properties": {}, "required": []}}},

    {"type": "function", "function": {
        "name": "browser_screenshot",
        "description": "Take a full-page screenshot and save as PNG.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string"}},
            "required": ["path"]}}},

    {"type": "function", "function": {
        "name": "browser_get_links",
        "description": "Return a list of links from the current page (max 40).",
        "parameters": {"type": "object", "properties": {}, "required": []}}},

    {"type": "function", "function": {
        "name": "browser_scroll",
        "description": "Scroll the page: up, down, top, or bottom.",
        "parameters": {"type": "object", "properties": {
            "direction": {"type": "string", "enum": ["up", "down", "top", "bottom"]}},
            "required": ["direction"]}}},

    {"type": "function", "function": {
        "name": "browser_press_key",
        "description": "Press a keyboard key: Enter, Tab, Escape, ArrowDown, etc.",
        "parameters": {"type": "object", "properties": {
            "key": {"type": "string"}},
            "required": ["key"]}}},

    {"type": "function", "function": {
        "name": "browser_select_option",
        "description": "Select an option from a <select> dropdown.",
        "parameters": {"type": "object", "properties": {
            "selector": {"type": "string"},
            "value":    {"type": "string"}},
            "required": ["selector", "value"]}}},

    {"type": "function", "function": {
        "name": "browser_wait",
        "description": "Wait for N seconds (max 30).",
        "parameters": {"type": "object", "properties": {
            "seconds": {"type": "number"}},
            "required": ["seconds"]}}},

    {"type": "function", "function": {
        "name": "browser_current_url",
        "description": "Return the current URL and page title.",
        "parameters": {"type": "object", "properties": {}, "required": []}}},

    {"type": "function", "function": {
        "name": "browser_go_back",
        "description": "Navigate back to the previous page.",
        "parameters": {"type": "object", "properties": {}, "required": []}}},

    {"type": "function", "function": {
        "name": "browser_eval_js",
        "description": "Execute JavaScript on the page and return the result.",
        "parameters": {"type": "object", "properties": {
            "script": {"type": "string"}},
            "required": ["script"]}}},

    # ── WEB (no browser) ────────────────────────────────────────────────────
    {"type": "function", "function": {
        "name": "read_webpage",
        "description": "Quickly fetch text content from a URL via HTTP (no browser, max 8000 chars).",
        "parameters": {"type": "object", "properties": {
            "url": {"type": "string"}},
            "required": ["url"]}}},

    # ── EXCEL ───────────────────────────────────────────────────────────────
    {"type": "function", "function": {
        "name": "create_excel",
        "description": (
            "Create a new Excel (.xlsx) file with data, headers, and auto-formatting. "
            "Use this to create spreadsheets from scratch."
        ),
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Path to the .xlsx file"},
            "sheets_data": {
                "type": "array",
                "description": "List of sheets to create",
                "items": {"type": "object", "properties": {
                    "name":       {"type": "string",  "description": "Sheet name"},
                    "headers":    {"type": "array",   "items": {"type": "string"}, "description": "Column headers"},
                    "rows":       {"type": "array",   "items": {"type": "array"},  "description": "Data rows"},
                    "col_widths": {"type": "array",   "items": {"type": "number"}, "description": "Column widths (optional)"}
                }}}},
            "required": ["path", "sheets_data"]}}},

    {"type": "function", "function": {
        "name": "read_excel",
        "description": "Read the contents of an Excel file (max 100 rows per sheet).",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string"}},
            "required": ["path"]}}},

    {"type": "function", "function": {
        "name": "edit_excel_cell",
        "description": "Edit the value of a single cell in an Excel file.",
        "parameters": {"type": "object", "properties": {
            "path":       {"type": "string"},
            "sheet_name": {"type": "string"},
            "cell":       {"type": "string", "description": "Cell address e.g. A1, B3"},
            "value":      {"description": "New cell value"}},
            "required": ["path", "sheet_name", "cell", "value"]}}},

    {"type": "function", "function": {
        "name": "add_excel_formula",
        "description": "Insert an Excel formula: =SUM(), =VLOOKUP(), =COUNTIF(), =IF(), =AVERAGE(), =MAX(), =MIN(), etc.",
        "parameters": {"type": "object", "properties": {
            "path":       {"type": "string"},
            "sheet_name": {"type": "string"},
            "cell":       {"type": "string"},
            "formula":    {"type": "string", "description": "Excel formula e.g. =SUM(D2:D11)"}},
            "required": ["path", "sheet_name", "cell", "formula"]}}},

    {"type": "function", "function": {
        "name": "add_excel_chart",
        "description": "Add a chart (bar / line / pie) to an Excel sheet.",
        "parameters": {"type": "object", "properties": {
            "path":       {"type": "string"},
            "sheet_name": {"type": "string"},
            "chart_type": {"type": "string", "enum": ["bar", "line", "pie"]},
            "data_range": {"type": "string", "description": "Data range e.g. A1:D10"},
            "title":      {"type": "string"},
            "position":   {"type": "string", "description": "Cell where chart is inserted e.g. F1"}},
            "required": ["path", "sheet_name", "chart_type", "data_range", "title", "position"]}}},

    {"type": "function", "function": {
        "name": "add_excel_sheet",
        "description": "Add a new sheet to an existing Excel file.",
        "parameters": {"type": "object", "properties": {
            "path":       {"type": "string"},
            "sheet_name": {"type": "string"}},
            "required": ["path", "sheet_name"]}}},

    {"type": "function", "function": {
        "name": "excel_add_rows",
        "description": "Append rows to the end of an Excel sheet.",
        "parameters": {"type": "object", "properties": {
            "path":       {"type": "string"},
            "sheet_name": {"type": "string"},
            "rows": {"type": "array", "items": {"type": "array"}}},
            "required": ["path", "sheet_name", "rows"]}}},

    {"type": "function", "function": {
        "name": "excel_style_range",
        "description": "Style a range of cells (bold, background color, font size).",
        "parameters": {"type": "object", "properties": {
            "path":       {"type": "string"},
            "sheet_name": {"type": "string"},
            "cell_range": {"type": "string"},
            "bold":       {"type": "boolean"},
            "bg_color":   {"type": "string", "description": "Hex color without # e.g. FF0000"},
            "font_size":  {"type": "integer"}},
            "required": ["path", "sheet_name", "cell_range"]}}},

    # ── SYSTEM ──────────────────────────────────────────────────────────────
    {"type": "function", "function": {
        "name": "run_command",
        "description": "Run a system command (CMD / PowerShell) and return the output.",
        "parameters": {"type": "object", "properties": {
            "command": {"type": "string"}},
            "required": ["command"]}}},
]

# ─────────────────────────────────────────
# TOOL DISPATCHER
# ─────────────────────────────────────────
TOOL_MAP = {
    "read_file":          lambda a: read_file(a["path"]),
    "write_file":         lambda a: write_file(a["path"], a["content"]),
    "list_files":         lambda a: list_files(a.get("directory", ".")),
    "open_file":          lambda a: open_file(a["path"]),
    "delete_file":        lambda a: delete_file(a["path"]),
    "copy_file":          lambda a: copy_file(a["src"], a["dst"]),
    "move_file":          lambda a: move_file(a["src"], a["dst"]),
    "create_directory":   lambda a: create_directory(a["path"]),
    "browser_goto":       lambda a: browser_goto(a["url"]),
    "browser_click":      lambda a: browser_click(a["selector"]),
    "browser_type":       lambda a: browser_type(a["selector"], a["text"]),
    "browser_get_text":   lambda a: browser_get_text(),
    "browser_screenshot": lambda a: browser_screenshot(a["path"]),
    "browser_get_links":  lambda a: browser_get_links(),
    "browser_scroll":     lambda a: browser_scroll(a["direction"]),
    "browser_press_key":  lambda a: browser_press_key(a["key"]),
    "browser_select_option": lambda a: browser_select_option(a["selector"], a["value"]),
    "browser_wait":       lambda a: browser_wait(a["seconds"]),
    "browser_current_url":lambda a: browser_current_url(),
    "browser_go_back":    lambda a: browser_go_back(),
    "browser_eval_js":    lambda a: browser_eval_js(a["script"]),
    "read_webpage":       lambda a: read_webpage(a["url"]),
    "create_excel":       lambda a: create_excel(a["path"], a["sheets_data"]),
    "read_excel":         lambda a: read_excel(a["path"]),
    "edit_excel_cell":    lambda a: edit_excel_cell(a["path"], a["sheet_name"], a["cell"], a["value"]),
    "add_excel_formula":  lambda a: add_excel_formula(a["path"], a["sheet_name"], a["cell"], a["formula"]),
    "add_excel_chart":    lambda a: add_excel_chart(a["path"], a["sheet_name"], a["chart_type"],
                                                     a["data_range"], a["title"], a["position"]),
    "add_excel_sheet":    lambda a: add_excel_sheet(a["path"], a["sheet_name"]),
    "excel_add_rows":     lambda a: excel_add_rows(a["path"], a["sheet_name"], a["rows"]),
    "excel_style_range":  lambda a: excel_style_range(a["path"], a["sheet_name"], a["cell_range"],
                                                       a.get("bold", False), a.get("bg_color"),
                                                       a.get("font_size")),
    "run_command":        lambda a: run_command(a["command"]),
}


def handle_tool_call(name: str, args: dict) -> str:
    preview = ", ".join(f"{k}={repr(v)[:50]}" for k, v in args.items())
    print(f"  [🔧 {name}({preview})]")
    handler = TOOL_MAP.get(name)
    if handler:
        return handler(args)
    return f"Unknown tool: {name}"


# ─────────────────────────────────────────
# API WITH RETRY AND AUTO-FALLBACK
# ─────────────────────────────────────────
def call_api(messages: list, use_tools: bool = True, model: str = None, retries: int = 3):
    """Call Groq API with automatic model fallback on errors."""
    if model is None:
        user_msgs = [m for m in messages if m["role"] == "user"]
        last_user = user_msgs[-1]["content"] if user_msgs else ""
        model     = choose_model(last_user)

    for attempt in range(retries):
        try:
            kwargs = {
                "model":      model,
                "messages":   messages,
                "max_tokens": 4096,
            }
            if use_tools:
                kwargs["tools"]       = tools
                kwargs["tool_choice"] = "auto"

            print(f"  [🤖 {model} | tools: {'✅' if use_tools else '❌'}]", end="", flush=True)
            response = client.chat.completions.create(**kwargs)
            print(" ✓")
            return response

        except Exception as e:
            error_str = str(e).lower()
            print(f"\n  [⚠️ Attempt {attempt+1}/{retries}: {str(e)[:80]}]")

            # Rate limit
            if "rate_limit" in error_str or "429" in error_str or "too many" in error_str:
                if model == MODEL_SMART:
                    print(f"  [↩️ Switching to {MODEL_FAST}]")
                    model = MODEL_FAST
                    continue
                wait_time = min(2 ** attempt * 2, 30)
                print(f"  [⏳ Waiting {wait_time}s...]")
                time.sleep(wait_time)
                continue

            # Tool use failed (only on this specific error, not generic "tool" string)
            if "tool_use_failed" in error_str:
                if attempt < retries - 1:
                    time.sleep(1)
                    continue

            # Model unavailable
            if "model" in error_str and ("not found" in error_str or "unavailable" in error_str):
                if model == MODEL_SMART:
                    print(f"  [↩️ {MODEL_SMART} unavailable, using {MODEL_FAST}]")
                    model = MODEL_FAST
                    continue

            if attempt < retries - 1:
                time.sleep(2)
                continue
            raise


def append_assistant(messages: list, msg) -> None:
    """Append assistant response to conversation history."""
    entry = {"role": "assistant", "content": msg.content or ""}
    if msg.tool_calls:
        entry["tool_calls"] = [
            {
                "id":   tc.id,
                "type": "function",
                "function": {
                    "name":      tc.function.name,
                    "arguments": tc.function.arguments,
                },
            }
            for tc in msg.tool_calls
        ]
    messages.append(entry)


# ─────────────────────────────────────────
# HISTORY MANAGEMENT
# ─────────────────────────────────────────
MAX_MESSAGES = 50


def trim_history(messages: list) -> list:
    """Trim history to prevent context overflow."""
    if len(messages) <= MAX_MESSAGES:
        return messages
    system  = [m for m in messages if m["role"] == "system"]
    rest    = [m for m in messages if m["role"] != "system"]
    trimmed = rest[-(MAX_MESSAGES - len(system)):]
    # Don't start with a tool response without a preceding tool call
    while trimmed and trimmed[0]["role"] == "tool":
        trimmed.pop(0)
    return system + trimmed


# ─────────────────────────────────────────
# SYSTEM PROMPT
# ─────────────────────────────────────────
SYSTEM_PROMPT = f"""You are an advanced AI assistant with full access to a Windows 11 computer.

AVAILABLE TOOLS (ALWAYS USE THEM when a task requires it):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📁 FILES: read_file, write_file, list_files, open_file, delete_file, copy_file, move_file, create_directory
🌐 BROWSER: browser_goto, browser_click, browser_type, browser_get_text, browser_screenshot, browser_get_links, browser_scroll, browser_press_key, browser_wait, browser_current_url, browser_go_back, browser_eval_js
🔗 WEB: read_webpage (fast HTTP fetch without browser)
📊 EXCEL: create_excel, read_excel, edit_excel_cell, add_excel_formula, add_excel_chart, add_excel_sheet, excel_add_rows, excel_style_range
⚙️ SYSTEM: run_command

CRITICAL RULES:
1. ALWAYS use tools — never say "I can't" or "the function is unavailable". You have access to ALL tools listed above.
2. Execute multi-step tasks autonomously without asking for confirmation at each step.
3. Briefly describe what you're doing at each step.
4. Use full Windows paths with backslashes or the Desktop path: {DESKTOP}
5. If something fails, try an alternative approach — never give up.
6. To search Google: browser_goto("google.com") → browser_type("q", "query") → browser_press_key("Enter").
7. To use Wikipedia: browser_goto("wikipedia.org") → search → browser_get_text().
8. Do NOT ask the user for data you can find yourself using tools.

TOOL REMINDERS:
- create_excel   → creates a new .xlsx file with data in one call
- add_excel_formula → adds formulas (=SUM, =MAX, =COUNTIF...) to an existing file
- add_excel_chart   → adds charts (bar/line/pie) to an existing file
- write_file        → creates any text file (txt, html, csv...)
- run_command       → runs CMD/PowerShell commands

Desktop path: {DESKTOP}
"""

# ─────────────────────────────────────────
# MAIN LOOP
# ─────────────────────────────────────────
messages = [{"role": "system", "content": SYSTEM_PROMPT}]

print()
print("╔══════════════════════════════════════════════════════╗")
print("║      🤖  GroqAgent  —  Autonomous AI for Windows    ║")
print("╠══════════════════════════════════════════════════════╣")
print(f"║  Smart : {MODEL_SMART:<42} ║")
print(f"║  Fast  : {MODEL_FAST:<42} ║")
print("╚══════════════════════════════════════════════════════╝")
print()
print(f"  📁 Desktop    : {DESKTOP}")
print(f"  🌐 Playwright : {'✅ OK' if PLAYWRIGHT_AVAILABLE else '❌ Missing  →  pip install playwright && playwright install chromium'}")
print(f"  📊 Excel      : {'✅ OK' if EXCEL_AVAILABLE     else '❌ Missing  →  pip install openpyxl'}")
print()
print("  Examples:")
print("  ──────────────────────────────────────────────────────")
print("  • Open google.com and search for the weather in London")
print("  • Take a screenshot of bbc.com and save to desktop")
print("  • Create an Excel budget spreadsheet with charts")
print("  • List all files on the desktop")
print("  • Run command: ipconfig")
print()
print("  Commands: 'exit' = quit  |  'reset' = clear history  |  'status' = stats")
print()

while True:
    try:
        user_input = input("👤 You: ").strip()
    except (KeyboardInterrupt, EOFError):
        print("\n\n👋 Goodbye!")
        close_browser()
        break

    if user_input.lower() in ("exit", "quit"):
        print("👋 Goodbye!")
        close_browser()
        break

    if user_input.lower() in ("reset", "clear"):
        messages = [messages[0]]
        print("🔄 History cleared.\n")
        continue

    if user_input.lower() in ("status", "stats"):
        print(f"\n📊 Status:")
        print(f"  Smart calls today  : {_smart_calls_today}/{_MAX_SMART_CALLS}")
        print(f"  Messages in history: {len(messages)}")
        print(f"  Browser            : {'open (' + _page.url + ')' if _page else 'closed'}\n")
        continue

    if not user_input:
        continue

    messages.append({"role": "user", "content": user_input})
    messages = trim_history(messages)

    try:
        MAX_ITERATIONS = 25
        iteration      = 0

        while iteration < MAX_ITERATIONS:
            iteration += 1
            response      = call_api(messages)
            msg           = response.choices[0].message
            finish_reason = response.choices[0].finish_reason

            append_assistant(messages, msg)

            # Model wants to use tools
            if msg.tool_calls:
                for tc in msg.tool_calls:
                    try:
                        args = json.loads(tc.function.arguments)
                    except json.JSONDecodeError:
                        args = {}

                    result     = handle_tool_call(tc.function.name, args)
                    result_str = str(result)
                    if len(result_str) > 8000:
                        result_str = result_str[:8000] + "\n[... truncated]"

                    messages.append({
                        "role":         "tool",
                        "tool_call_id": tc.id,
                        "content":      result_str,
                    })
                continue  # loop back so model processes tool results

            # Model replied with text
            if msg.content:
                print(f"\n🤖 Agent: {msg.content}\n")
            break

        else:
            print(f"\n⚠️ Reached iteration limit ({MAX_ITERATIONS}).\n")

    except Exception as e:
        print(f"\n❌ Error: {e}\n")
        # Remove last user message to avoid corrupting history
        if messages and messages[-1]["role"] == "user":
            messages.pop()
