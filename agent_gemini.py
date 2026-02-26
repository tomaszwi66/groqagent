# ─────────────────────────────────────────
# ENCODING - MUST BE FIRST
# ─────────────────────────────────────────
import sys
import os
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')
os.environ["PYTHONIOENCODING"] = "utf-8"

# ─────────────────────────────────────────
# STANDARD IMPORTS
# ─────────────────────────────────────────
import json
import time
import urllib.request
from html.parser import HTMLParser
from datetime import datetime, date

# ─────────────────────────────────────────
# GOOGLE GEMINI - AI CLIENT
# ─────────────────────────────────────────
from google import genai
from google.genai import types

API_KEY = ""  # ← PASTE YOUR GOOGLE AI STUDIO KEY HERE
client  = genai.Client(api_key=API_KEY)

# ─────────────────────────────────────────
# MODELS - DUAL STRATEGY
# ─────────────────────────────────────────
MODEL_FAST  = "gemini-2.0-flash"  # fast, generous free tier
MODEL_SMART = "gemini-2.5-pro"    # most capable

_smart_calls_today = 0
_smart_calls_date  = date.today()
_MAX_SMART_CALLS   = 100  # AI Studio free tier safety buffer


def choose_model(user_message: str) -> str:
    """Automatically selects model based on task complexity."""
    global _smart_calls_today, _smart_calls_date
    if date.today() != _smart_calls_date:
        _smart_calls_today = 0
        _smart_calls_date  = date.today()

    msg_lower = user_message.lower()
    simple_keywords  = ["open","click","type","screenshot","save","read","scroll",
                        "wait","close","show","list","go to","navigate","press"]
    complex_keywords = ["analyz","compare","strategy","plan","summarize","summary",
                        "report","conclusions","optimize","explain","why","create",
                        "design","budget","calculate","step by step","following",
                        "perform","excel","spreadsheet","html","table","chart","graph"]

    is_simple  = any(kw in msg_lower for kw in simple_keywords)
    is_complex = any(kw in msg_lower for kw in complex_keywords) or len(user_message) > 200

    if is_complex and _smart_calls_today < _MAX_SMART_CALLS:
        _smart_calls_today += 1
        return MODEL_SMART
    elif is_simple and not is_complex:
        return MODEL_FAST
    elif _smart_calls_today < _MAX_SMART_CALLS:
        _smart_calls_today += 1
        return MODEL_SMART
    return MODEL_FAST


# ─────────────────────────────────────────
# PLAYWRIGHT
# ─────────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("[⚠️  Playwright not available. Run: pip install playwright && playwright install chromium]")

# ─────────────────────────────────────────
# OPENPYXL (EXCEL)
# ─────────────────────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[⚠️  openpyxl not available. Run: pip install openpyxl]")

# ─────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")

_playwright = None
_browser    = None
_page       = None


def fix_path(path: str) -> str:
    return (path
            .replace("~/Desktop", DESKTOP)
            .replace("/desktop",  DESKTOP)
            .replace("~/desktop", DESKTOP))


# ─────────────────────────────────────────
# BROWSER
# ─────────────────────────────────────────
def get_page():
    global _playwright, _browser, _page
    if _page is None:
        _playwright = sync_playwright().start()
        _browser    = _playwright.chromium.launch(headless=False, slow_mo=100)
        _page       = _browser.new_page()
        _page.set_viewport_size({"width": 1280, "height": 800})
    return _page


def close_browser():
    global _playwright, _browser, _page
    if _browser:
        try: _browser.close()
        except: pass
    if _playwright:
        try: _playwright.stop()
        except: pass
    _page = _browser = _playwright = None


def browser_goto(url: str) -> str:
    """Navigate to a URL in the Chromium browser."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        if not url.startswith("http"): url = "https://" + url
        page = get_page()
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        return f"Opened: {url} | Title: {page.title()}"
    except Exception as e:
        return f"Navigation error: {e}"


def browser_click(selector: str) -> str:
    """Click an element on the page by visible text or CSS selector."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        page = get_page(); clicked = False
        for s in [
            lambda: page.get_by_text(selector, exact=False).first.click(timeout=3000),
            lambda: page.get_by_role("button", name=selector).first.click(timeout=3000),
            lambda: page.get_by_role("link",   name=selector).first.click(timeout=3000),
            lambda: page.click(selector, timeout=3000),
        ]:
            try: s(); clicked = True; break
            except: pass
        if not clicked: return f"Element not found: {selector}"
        page.wait_for_load_state("domcontentloaded")
        return f"Clicked: {selector}"
    except Exception as e:
        return f"Click error: {e}"


def browser_type(selector: str, text: str) -> str:
    """Type text into a form field identified by placeholder, label, or CSS selector."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        page = get_page(); filled = False
        for s in [
            lambda: page.get_by_placeholder(selector).first.fill(text, timeout=3000),
            lambda: page.get_by_label(selector).first.fill(text, timeout=3000),
            lambda: page.get_by_role("textbox",   name=selector).first.fill(text, timeout=3000),
            lambda: page.get_by_role("searchbox").first.fill(text, timeout=3000),
            lambda: page.fill(selector, text, timeout=3000),
        ]:
            try: s(); filled = True; break
            except: pass
        if not filled: return f"Field not found: {selector}"
        return f"Typed '{text}' into: {selector}"
    except Exception as e:
        return f"Type error: {e}"


def browser_get_text() -> str:
    """Get all visible text from the current page (max 6000 chars)."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        text = get_page().inner_text("body")
        return (text[:6000] + "\n[... truncated]") if len(text) > 6000 else text
    except Exception as e:
        return f"Get text error: {e}"


def browser_screenshot(path: str) -> str:
    """Take a full-page screenshot and save as PNG."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    path = fix_path(path)
    if not path.endswith(".png"): path += ".png"
    try:
        d = os.path.dirname(path)
        if d: os.makedirs(d, exist_ok=True)
        get_page().screenshot(path=path, full_page=True)
        return f"Screenshot saved: {path}"
    except Exception as e:
        return f"Screenshot error: {e}"


def browser_get_links() -> str:
    """Return a list of all links on the current page (max 40)."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        links = get_page().eval_on_selector_all(
            "a[href]",
            "els => els.map(e=>({text:e.innerText.trim(),href:e.href})).filter(l=>l.text&&l.href)")
        return "\n".join(f"{l['text'][:60]} -> {l['href']}" for l in links[:40]) or "No links."
    except Exception as e:
        return f"Get links error: {e}"


def browser_scroll(direction: str) -> str:
    """Scroll the page: up, down, top, or bottom."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        get_page().keyboard.press(
            {"down":"PageDown","up":"PageUp","top":"Home","bottom":"End"}.get(direction,"PageDown"))
        time.sleep(0.3); return f"Scrolled: {direction}"
    except Exception as e:
        return f"Scroll error: {e}"


def browser_press_key(key: str) -> str:
    """Press a keyboard key: Enter, Tab, Escape, ArrowDown, etc."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        get_page().keyboard.press(key); time.sleep(0.5); return f"Pressed: {key}"
    except Exception as e:
        return f"Key press error: {e}"


def browser_select_option(selector: str, value: str) -> str:
    """Select an option from a dropdown by label or value."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        page = get_page()
        try: page.select_option(selector, label=value, timeout=5000)
        except: page.select_option(selector, value=value, timeout=5000)
        return f"Selected '{value}' in: {selector}"
    except Exception as e:
        return f"Select error: {e}"


def browser_wait(seconds: float) -> str:
    """Wait a given number of seconds (max 30)."""
    try:
        time.sleep(min(float(seconds), 30)); return f"Waited {seconds}s"
    except Exception as e:
        return f"Wait error: {e}"


def browser_current_url() -> str:
    """Return the current URL and page title."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        p = get_page(); return f"URL: {p.url} | Title: {p.title()}"
    except Exception as e:
        return f"URL error: {e}"


def browser_go_back() -> str:
    """Go back to the previous page in the browser."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        get_page().go_back(wait_until="domcontentloaded", timeout=10000)
        return f"Went back. URL: {get_page().url}"
    except Exception as e:
        return f"Go back error: {e}"


def browser_eval_js(script: str) -> str:
    """Execute JavaScript on the page and return the result."""
    if not PLAYWRIGHT_AVAILABLE: return "Playwright not available."
    try:
        result = get_page().evaluate(script)
        return str(result)[:3000] if result else "OK (no result)"
    except Exception as e:
        return f"JS error: {e}"


# ─────────────────────────────────────────
# WEB FETCH (without browser)
# ─────────────────────────────────────────
class TextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.text = []; self.skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ("script","style","nav","footer","head","noscript"): self.skip = True

    def handle_endtag(self, tag):
        if tag in ("script","style","nav","footer","head","noscript"): self.skip = False

    def handle_data(self, data):
        if not self.skip:
            s = data.strip()
            if s: self.text.append(s)


def read_webpage(url: str) -> str:
    """Fast HTTP page text fetch without a browser (max 8000 chars)."""
    try:
        if not url.startswith("http"): url = "https://" + url
        req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as r:
            html = r.read().decode("utf-8", errors="ignore")
        parser = TextExtractor(); parser.feed(html)
        text = "\n".join(parser.text)
        return (text[:8000] + "\n[... truncated]") if len(text) > 8000 else text or "No content."
    except Exception as e:
        return f"Fetch error: {e}"


# ─────────────────────────────────────────
# FILES
# ─────────────────────────────────────────
def read_file(path: str) -> str:
    """Read the contents of a text file."""
    path = fix_path(path)
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        return (content[:10000] + "\n[... truncated]") if len(content) > 10000 else content
    except UnicodeDecodeError:
        try:
            with open(path, "r", encoding="cp1250") as f: return f.read()[:10000]
        except Exception as e: return f"Read error (encoding): {e}"
    except Exception as e:
        return f"Read error: {e}"


def write_file(path: str, content: str) -> str:
    """Write text to a file. Creates parent directories if needed."""
    path = fix_path(path)
    try:
        d = os.path.dirname(path)
        if d: os.makedirs(d, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f: f.write(content)
        return f"Saved: {path} ({len(content)} chars)"
    except Exception as e:
        return f"Write error: {e}"


def list_files(directory: str = ".") -> str:
    """List files and folders in a directory with sizes."""
    directory = fix_path(directory)
    try:
        result = []
        for entry in sorted(os.listdir(directory)):
            full = os.path.join(directory, entry)
            if os.path.isdir(full):
                result.append(f"📁 {entry}/")
            else:
                size = os.path.getsize(full)
                s = f"{size} B" if size < 1024 else (f"{size/1024:.1f} KB" if size < 1024**2 else f"{size/1024/1024:.1f} MB")
                result.append(f"📄 {entry} ({s})")
        return "\n".join(result) if result else "Directory is empty."
    except Exception as e:
        return f"List error: {e}"


def open_file(path: str) -> str:
    """Open a file in its default Windows application."""
    path = fix_path(path)
    try:
        os.startfile(path); return f"Opened: {path}"
    except Exception as e:
        return f"Open error: {e}"


def delete_file(path: str) -> str:
    """Delete a file or folder."""
    path = fix_path(path)
    try:
        if os.path.isfile(path):
            os.remove(path); return f"Deleted file: {path}"
        elif os.path.isdir(path):
            import shutil; shutil.rmtree(path); return f"Deleted folder: {path}"
        return f"Not found: {path}"
    except Exception as e:
        return f"Delete error: {e}"


def copy_file(src: str, dst: str) -> str:
    """Copy a file to a new location."""
    src, dst = fix_path(src), fix_path(dst)
    try:
        import shutil
        d = os.path.dirname(dst)
        if d: os.makedirs(d, exist_ok=True)
        shutil.copy2(src, dst); return f"Copied: {src} -> {dst}"
    except Exception as e:
        return f"Copy error: {e}"


def move_file(src: str, dst: str) -> str:
    """Move or rename a file."""
    src, dst = fix_path(src), fix_path(dst)
    try:
        import shutil
        d = os.path.dirname(dst)
        if d: os.makedirs(d, exist_ok=True)
        shutil.move(src, dst); return f"Moved: {src} -> {dst}"
    except Exception as e:
        return f"Move error: {e}"


def create_directory(path: str) -> str:
    """Create a folder recursively."""
    path = fix_path(path)
    try:
        os.makedirs(path, exist_ok=True); return f"Created folder: {path}"
    except Exception as e:
        return f"mkdir error: {e}"


# ─────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────
def _thin_border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _create_excel_impl(path: str, sheets_data: list) -> str:
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    if not path.endswith(".xlsx"): path += ".xlsx"
    try:
        wb = Workbook(); wb.remove(wb.active)
        for sd in sheets_data:
            ws = wb.create_sheet(title=sd.get("name", "Sheet1"))
            headers = sd.get("headers", []); rows = sd.get("rows", [])
            if headers:
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _thin_border()
            for row in rows:
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.border = _thin_border()
                    cell.alignment = Alignment(vertical="center")
            col_widths = sd.get("col_widths", [])
            if col_widths:
                for i, w in enumerate(col_widths):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = w
            else:
                for col in ws.columns:
                    mx = max((len(str(c.value)) for c in col if c.value), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(mx+4, 50)
        d = os.path.dirname(path)
        if d: os.makedirs(d, exist_ok=True)
        wb.save(path); return f"Excel '{path}' created ({len(sheets_data)} sheet(s))."
    except Exception as e:
        return f"Excel create error: {e}"


def read_excel(path: str) -> str:
    """Read the contents of an Excel file (max 100 rows per sheet)."""
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path, data_only=True); result = []
        for sn in wb.sheetnames:
            ws = wb[sn]; result.append(f"=== Sheet: {sn} ({ws.max_row}r x {ws.max_column}c) ===")
            n = 0
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    result.append("\t".join(str(c) if c is not None else "" for c in row))
                    n += 1
                    if n > 100: result.append("[... truncated]"); break
        return "\n".join(result) or "File is empty."
    except Exception as e:
        return f"Excel read error: {e}"


def edit_excel_cell(path: str, sheet_name: str, cell: str, value: str) -> str:
    """Edit the value of a single cell in an Excel file."""
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path); ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        try:
            v = float(value); value = int(v) if v == int(v) else v
        except (ValueError, TypeError): pass
        ws[cell] = value; wb.save(path); return f"Cell {sheet_name}!{cell} = {value}"
    except Exception as e:
        return f"Cell edit error: {e}"


def add_excel_formula(path: str, sheet_name: str, cell: str, formula: str) -> str:
    """Insert an Excel formula like =SUM(), =VLOOKUP(), =COUNTIF(), =IF(), =AVERAGE(), =MAX(), =MIN()."""
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path); ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws[cell] = formula; wb.save(path); return f"Formula '{formula}' set in {sheet_name}!{cell}"
    except Exception as e:
        return f"Formula error: {e}"


def add_excel_chart(path: str, sheet_name: str, chart_type: str,
                    data_range: str, title: str, position: str) -> str:
    """Add a chart (bar, line, or pie) to an Excel sheet."""
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path); ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ref = Reference(ws, range_string=f"{sheet_name}!{data_range}")
        chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}.get(chart_type, BarChart)()
        chart.title = title; chart.style = 10; chart.width = 18; chart.height = 12
        chart.add_data(ref, titles_from_data=True); ws.add_chart(chart, position)
        wb.save(path); return f"Chart '{chart_type}' '{title}' added at {position}."
    except Exception as e:
        return f"Chart error: {e}"


def add_excel_sheet(path: str, sheet_name: str) -> str:
    """Add a new sheet to an existing Excel file."""
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames: wb.create_sheet(sheet_name)
        wb.save(path); return f"Sheet '{sheet_name}' added."
    except Exception as e:
        return f"Add sheet error: {e}"


def _excel_add_rows_impl(path: str, sheet_name: str, rows: list) -> str:
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path); ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in rows:
            ws.append(row)
            for cell in ws[ws.max_row]: cell.border = _thin_border()
        wb.save(path); return f"Added {len(rows)} row(s) to '{sheet_name}'."
    except Exception as e:
        return f"Add rows error: {e}"


def excel_style_range(path: str, sheet_name: str, cell_range: str,
                      bold: bool = False, bg_color: str = None, font_size: int = None) -> str:
    """Style a range of cells with bold, background color, or font size."""
    if not EXCEL_AVAILABLE: return "openpyxl not available."
    path = fix_path(path)
    try:
        wb = load_workbook(path); ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in ws[cell_range]:
            if not isinstance(row, tuple): row = (row,)
            for cell in row:
                if bold or font_size: cell.font = Font(bold=bold, size=font_size or cell.font.size)
                if bg_color: cell.fill = PatternFill("solid", fgColor=bg_color)
        wb.save(path); return f"Style applied to {cell_range}."
    except Exception as e:
        return f"Style error: {e}"


# ─────────────────────────────────────────
# SYSTEM COMMANDS
# ─────────────────────────────────────────
def run_command(command: str) -> str:
    """Run a CMD or PowerShell command and return the output."""
    try:
        import subprocess
        result = subprocess.run(command, shell=True, capture_output=True, text=True,
                                timeout=30, encoding="utf-8", errors="ignore")
        output = result.stdout + result.stderr
        return (output[:5000] + "\n[... truncated]") if len(output) > 5000 else output or "Command executed (no output)."
    except subprocess.TimeoutExpired:
        return "Timeout — command took too long."
    except Exception as e:
        return f"Command error: {e}"


# ─────────────────────────────────────────
# TOOL DISPATCHER
# Maps tool name → Python function
# ─────────────────────────────────────────
TOOL_MAP = {
    # files
    "read_file":             lambda a: read_file(a["path"]),
    "write_file":            lambda a: write_file(a["path"], a["content"]),
    "list_files":            lambda a: list_files(a.get("directory", ".")),
    "open_file":             lambda a: open_file(a["path"]),
    "delete_file":           lambda a: delete_file(a["path"]),
    "copy_file":             lambda a: copy_file(a["src"], a["dst"]),
    "move_file":             lambda a: move_file(a["src"], a["dst"]),
    "create_directory":      lambda a: create_directory(a["path"]),
    # browser
    "browser_goto":          lambda a: browser_goto(a["url"]),
    "browser_click":         lambda a: browser_click(a["selector"]),
    "browser_type":          lambda a: browser_type(a["selector"], a["text"]),
    "browser_get_text":      lambda a: browser_get_text(),
    "browser_screenshot":    lambda a: browser_screenshot(a["path"]),
    "browser_get_links":     lambda a: browser_get_links(),
    "browser_scroll":        lambda a: browser_scroll(a["direction"]),
    "browser_press_key":     lambda a: browser_press_key(a["key"]),
    "browser_select_option": lambda a: browser_select_option(a["selector"], a["value"]),
    "browser_wait":          lambda a: browser_wait(a["seconds"]),
    "browser_current_url":   lambda a: browser_current_url(),
    "browser_go_back":       lambda a: browser_go_back(),
    "browser_eval_js":       lambda a: browser_eval_js(a["script"]),
    # web
    "read_webpage":          lambda a: read_webpage(a["url"]),
    # excel
    "create_excel":          lambda a: _create_excel_impl(a["path"], a["sheets_data"]),
    "read_excel":            lambda a: read_excel(a["path"]),
    "edit_excel_cell":       lambda a: edit_excel_cell(a["path"], a["sheet_name"], a["cell"], a["value"]),
    "add_excel_formula":     lambda a: add_excel_formula(a["path"], a["sheet_name"], a["cell"], a["formula"]),
    "add_excel_chart":       lambda a: add_excel_chart(a["path"], a["sheet_name"], a["chart_type"],
                                                        a["data_range"], a["title"], a["position"]),
    "add_excel_sheet":       lambda a: add_excel_sheet(a["path"], a["sheet_name"]),
    "excel_add_rows":        lambda a: _excel_add_rows_impl(a["path"], a["sheet_name"], a["rows"]),
    "excel_style_range":     lambda a: excel_style_range(a["path"], a["sheet_name"], a["cell_range"],
                                                          a.get("bold", False), a.get("bg_color"),
                                                          a.get("font_size")),
    # system
    "run_command":           lambda a: run_command(a["command"]),
}


def handle_tool_call(name: str, args: dict) -> str:
    preview = ", ".join(f"{k}={repr(v)[:50]}" for k, v in args.items())
    print(f"  [🔧 {name}({preview})]")
    handler = TOOL_MAP.get(name)
    return str(handler(args)) if handler else f"Unknown tool: {name}"


# ─────────────────────────────────────────
# GEMINI TOOL DECLARATIONS
#
# Simple tools  → pass Python function directly
#   (SDK infers schema from docstring + type hints)
#
# Complex tools → use types.FunctionDeclaration
#   with explicit Schema for nested list params
#   (fixes INVALID_ARGUMENT for list-of-objects / list-of-lists)
# ─────────────────────────────────────────

# ── Schema helpers ──────────────────────
S = types.Schema

def _prop(**kwargs): return S(**kwargs)

# ── Simple tools (auto-schema from Python functions) ──
SIMPLE_TOOLS = [
    read_file, write_file, list_files, open_file, delete_file,
    copy_file, move_file, create_directory,
    browser_goto, browser_click, browser_type, browser_get_text,
    browser_screenshot, browser_get_links, browser_scroll,
    browser_press_key, browser_select_option, browser_wait,
    browser_current_url, browser_go_back, browser_eval_js,
    read_webpage,
    read_excel, edit_excel_cell, add_excel_formula,
    add_excel_chart, add_excel_sheet, excel_style_range,
    run_command,
]

# ── create_excel — manually declared (nested list-of-objects) ──
_sheet_schema = S(
    type=types.Type.OBJECT,
    properties={
        "name": S(type=types.Type.STRING, description="Sheet name"),
        "headers": S(
            type=types.Type.ARRAY,
            description="Column header names",
            items=S(type=types.Type.STRING),
        ),
        "rows": S(
            type=types.Type.ARRAY,
            description="Data rows; each row is a list of cell values",
            items=S(
                type=types.Type.ARRAY,
                items=S(type=types.Type.STRING),
            ),
        ),
        "col_widths": S(
            type=types.Type.ARRAY,
            description="Optional column widths",
            items=S(type=types.Type.NUMBER),
        ),
    },
)

_create_excel_decl = types.FunctionDeclaration(
    name="create_excel",
    description=(
        "Create a new Excel .xlsx file with data, headers, and auto-formatting. "
        "Use this to create spreadsheets from scratch."
    ),
    parameters=S(
        type=types.Type.OBJECT,
        properties={
            "path": S(type=types.Type.STRING, description="Full path to the .xlsx file"),
            "sheets_data": S(
                type=types.Type.ARRAY,
                description="List of sheets to create",
                items=_sheet_schema,
            ),
        },
        required=["path", "sheets_data"],
    ),
)

# ── excel_add_rows — manually declared (list-of-lists) ──
_excel_add_rows_decl = types.FunctionDeclaration(
    name="excel_add_rows",
    description="Append rows to the end of an Excel sheet.",
    parameters=S(
        type=types.Type.OBJECT,
        properties={
            "path":       S(type=types.Type.STRING, description="Path to the .xlsx file"),
            "sheet_name": S(type=types.Type.STRING, description="Name of the sheet"),
            "rows": S(
                type=types.Type.ARRAY,
                description="Rows to append; each row is a list of cell values",
                items=S(
                    type=types.Type.ARRAY,
                    items=S(type=types.Type.STRING),
                ),
            ),
        },
        required=["path", "sheet_name", "rows"],
    ),
)

# ── Final tool list passed to Gemini ──
GEMINI_TOOLS = types.Tool(function_declarations=[
    # Auto-schema tools — wrap each function in FunctionDeclaration via Tool
    *[f for f in SIMPLE_TOOLS],   # SDK accepts callables directly alongside declarations
    _create_excel_decl,
    _excel_add_rows_decl,
])


# ─────────────────────────────────────────
# SYSTEM PROMPT
# ─────────────────────────────────────────
SYSTEM_PROMPT = f"""You are an advanced AI assistant with full access to a Windows 11 computer.

AVAILABLE TOOLS (ALWAYS USE THEM when the task requires it):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📁 FILES: read_file, write_file, list_files, open_file, delete_file, copy_file, move_file, create_directory
🌐 BROWSER: browser_goto, browser_click, browser_type, browser_get_text, browser_screenshot, browser_get_links, browser_scroll, browser_press_key, browser_wait, browser_current_url, browser_go_back, browser_eval_js
🔗 WEB: read_webpage (fast HTTP fetch without a browser)
📊 EXCEL: create_excel, read_excel, edit_excel_cell, add_excel_formula, add_excel_chart, add_excel_sheet, excel_add_rows, excel_style_range
⚙️ SYSTEM: run_command

CRITICAL RULES:
1. ALWAYS use tools — never say "I can't" or "that function is unavailable".
2. Execute multi-step tasks autonomously without asking for confirmation at every step.
3. Briefly state what you are doing at each step.
4. Use full Windows paths with backslashes or the desktop path: {DESKTOP}
5. If something fails, try an alternative approach — don't give up.
6. For Google search: browser_goto("google.com") → browser_type("q", "query") → browser_press_key("Enter").
7. For Wikipedia: browser_goto("wikipedia.org") → search → browser_get_text().
8. Do NOT ask the user for data you can find yourself using tools.

TOOL REMINDERS:
- create_excel      → creates an Excel file with all data in one call.
- add_excel_formula → adds formulas to an existing file.
- add_excel_chart   → adds charts to an existing file.
- write_file        → creates any text, HTML, or report file.

User desktop: {DESKTOP}"""


# ─────────────────────────────────────────
# CONVERSATION HISTORY
# ─────────────────────────────────────────
MAX_TURNS = 50
history: list[types.Content] = []


def trim_history():
    global history
    if len(history) > MAX_TURNS * 2:
        history = history[-(MAX_TURNS * 2):]


# ─────────────────────────────────────────
# MAIN API CALL
# ─────────────────────────────────────────
def call_gemini(user_text: str, retries: int = 3) -> str:
    """Send a message and run the tool-use loop until the model replies with text."""
    global history

    model = choose_model(user_text)
    history.append(types.Content(role="user", parts=[types.Part(text=user_text)]))
    trim_history()

    config = types.GenerateContentConfig(
        system_instruction=SYSTEM_PROMPT,
        tools=[GEMINI_TOOLS],
        temperature=0.2,
        max_output_tokens=4096,
    )

    for attempt in range(retries):
        try:
            print(f"  [🤖 {model} | tools: ✅]", end="", flush=True)

            MAX_ITERATIONS = 25
            iteration      = 0

            while iteration < MAX_ITERATIONS:
                iteration += 1

                response      = client.models.generate_content(
                    model    = model,
                    contents = history,
                    config   = config,
                )
                print(" ✓")

                model_content = response.candidates[0].content
                history.append(model_content)

                tool_parts = [p for p in model_content.parts if p.function_call]
                text_parts = [p for p in model_content.parts if p.text]

                if not tool_parts:
                    return " ".join(p.text for p in text_parts if p.text).strip()

                # Execute tools and feed results back
                result_parts = []
                for part in tool_parts:
                    fc   = part.function_call
                    args = dict(fc.args) if fc.args else {}
                    result_parts.append(types.Part(
                        function_response=types.FunctionResponse(
                            name    = fc.name,
                            response= {"result": handle_tool_call(fc.name, args)},
                        )
                    ))

                history.append(types.Content(role="user", parts=result_parts))
                print(f"  [🤖 {model} | tools: ✅]", end="", flush=True)

            return "⚠️ Reached iteration limit."

        except Exception as e:
            error_str = str(e).lower()
            print(f"\n  [⚠️  Attempt {attempt+1}/{retries}: {str(e)[:120]}]")

            if "quota" in error_str or "429" in error_str or "resource_exhausted" in error_str:
                if model == MODEL_SMART:
                    print(f"  [↩️  Switching to {MODEL_FAST}]")
                    model = MODEL_FAST; continue
                wait_time = min(2 ** attempt * 5, 60)
                print(f"  [⏳ Waiting {wait_time}s...]")
                time.sleep(wait_time); continue

            if "not found" in error_str or "unavailable" in error_str:
                if model == MODEL_SMART:
                    print(f"  [↩️  Falling back to {MODEL_FAST}]")
                    model = MODEL_FAST; continue

            if attempt < retries - 1:
                time.sleep(3); continue
            raise


# ─────────────────────────────────────────
# STARTUP BANNER
# ─────────────────────────────────────────
print()
print("╔══════════════════════════════════════════════════════╗")
print("║      🤖  GeminiAgent  (Dual Model Strategy)         ║")
print("╠══════════════════════════════════════════════════════╣")
print(f"║  Smart : {MODEL_SMART:<42} ║")
print(f"║  Fast  : {MODEL_FAST:<42} ║")
print("╚══════════════════════════════════════════════════════╝")
print()
print(f"  📁 Desktop    : {DESKTOP}")
print(f"  🌐 Playwright : {'✅ OK' if PLAYWRIGHT_AVAILABLE else '❌ Missing  →  pip install playwright && playwright install chromium'}")
print(f"  📊 Excel      : {'✅ OK' if EXCEL_AVAILABLE     else '❌ Missing  →  pip install openpyxl'}")
print()
print("  Examples:")
print("  ──────────────────────────────────────────────────────")
print("  • Open google.com and search for the weather in London")
print("  • Take a screenshot of bbc.com and save to the desktop")
print("  • Create an Excel budget with charts")
print("  • List all files on the desktop")
print("  • Run command: ipconfig")
print()
print("  Commands: 'exit' = quit | 'reset' = clear history | 'status' = stats")
print()

# ─────────────────────────────────────────
# MAIN LOOP
# ─────────────────────────────────────────
while True:
    try:
        user_input = input("👤 You: ").strip()
    except (KeyboardInterrupt, EOFError):
        print("\n\n👋 Goodbye!")
        close_browser(); break

    if user_input.lower() in ("exit", "quit"):
        print("👋 Goodbye!")
        close_browser(); break

    if user_input.lower() in ("reset", "clear"):
        history = []
        print("🔄 History cleared.\n"); continue

    if user_input.lower() in ("status", "stats"):
        print(f"\n📊 Status:")
        print(f"  Smart calls today  : {_smart_calls_today}/{_MAX_SMART_CALLS}")
        print(f"  Turns in history   : {len(history) // 2}")
        print(f"  Browser            : {'open (' + _page.url + ')' if _page else 'closed'}\n")
        continue

    if not user_input:
        continue

    try:
        reply = call_gemini(user_input)
        if reply:
            print(f"\n🤖 Agent: {reply}\n")
    except Exception as e:
        print(f"\n❌ Error: {e}\n")
        if history and history[-1].role == "user":
            history.pop()
